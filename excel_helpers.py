import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference

from area_helpers import clean_area_records, clean_door_records


class ExcelImportError(Exception):
    """User-facing Excel import validation error."""
    pass


def _excel_norm_col(x):
    return (
        str(x)
        .strip()
        .lower()
        .replace("\n", " ")
        .replace("_", " ")
        .replace("-", " ")
        .replace("/", " ")
        .replace(".", "")
    )

def _excel_find_col(df, names):
    norm_map = {_excel_norm_col(col): col for col in df.columns}

    for name in names:
        key = _excel_norm_col(name)
        if key in norm_map:
            return norm_map[key]

    return None

def _excel_safe_float(v, default=0.0):
    if v is None:
        return default

    text = str(v).strip()

    if text in ["", "-", "None", "nan", "NaN"]:
        return default

    text = text.replace(",", "")

    try:
        return float(text)
    except Exception:
        return default

def _excel_safe_int(v, default=0):
    return int(_excel_safe_float(v, default))

def _read_excel_sheet_with_header(excel_bytes, sheet_name, header_candidates):
    """
    Prefer hidden stable import key row if available.
    This avoids reading the pretty merged header row as the real header.
    """
    try:
        raw = pd.read_excel(
            io.BytesIO(excel_bytes),
            sheet_name=sheet_name,
            header=None,
            engine="openpyxl",
        )
    except ValueError as e:
        if sheet_name in str(e):
            raise ExcelImportError(f'The workbook must contain an "{sheet_name}" sheet.') from e
        raise ExcelImportError(
            "Could not read the workbook. Please upload a valid .xlsx file generated from this app."
        ) from e
    except Exception as e:
        raise ExcelImportError(
            "Could not read the workbook. Please upload a valid .xlsx file generated from this app."
        ) from e

    header_row_idx = None

    # Prefer row that contains stable backend keys.
    stable_required = [
        "FL",
        "Space Type",
    ]

    stable_area_keys = [
        "Floor to Floor Height (m)",
        "Typical Unit",
        "Unit Area",
        "Stair, MEP, Etc",
        "Koridor/Lobby",
    ]

    for i in range(min(12, len(raw))):
        row_values = [str(v).strip() for v in raw.iloc[i].tolist()]
        row_norm = [_excel_norm_col(v) for v in row_values]

        stable_hits = sum(
            1 for key in stable_required + stable_area_keys
            if _excel_norm_col(key) in row_norm
        )

        # Hidden key row should hit many exact backend names.
        if stable_hits >= 4:
            header_row_idx = i
            break

    # Fallback to old detection
    if header_row_idx is None:
        for i in range(min(12, len(raw))):
            row_values = [
                _excel_norm_col(v)
                for v in raw.iloc[i].tolist()
                if str(v).strip() not in ["", "nan", "None"]
            ]

            joined = " | ".join(row_values)

            hits = 0
            for candidate in header_candidates:
                if _excel_norm_col(candidate) in joined:
                    hits += 1

            if hits >= 2:
                header_row_idx = i
                break

    if header_row_idx is None:
        raise ExcelImportError(
            "Could not find the header row. Please use the latest downloaded Excel template."
        )

    df = raw.iloc[header_row_idx + 1:].copy()
    df.columns = raw.iloc[header_row_idx].tolist()

    df = df.loc[:, [str(c).strip() not in ["", "nan", "None"] for c in df.columns]]
    df = df.dropna(how="all").reset_index(drop=True)

    if df.empty:
        raise ExcelImportError(
            f'The "{sheet_name}" sheet is empty after the header row. Please use the latest downloaded Excel template.'
        )

    # Remove accidental repeated header rows.
    if "FL" in df.columns:
        df = df[
            ~df["FL"].astype(str).str.strip().str.upper().isin(
                ["FL", "FLOOR", "TOTAL", "GRAND TOTAL"]
            )
        ].copy()

    return df

def read_area_input_sheet(excel_bytes):
    df = _read_excel_sheet_with_header(
        excel_bytes,
        "Area Input",
        header_candidates=["FL", "FLOOR", "Parkir", "GFA", "SGFA", "NFA"],
    )

    fl_col = _excel_find_col(df, ["FL", "Floor", "FLOOR"])
    space_col = _excel_find_col(df, ["Space Type", "Floor Type", "Type"])
    f2f_col = _excel_find_col(df, ["Floor to Floor Height (m)", "FL TO FL HEIGHT", "Height", "Height (m)"])
    typical_col = _excel_find_col(df, ["Typical Unit", "Unit Count", "Typical Units"])

    parkir_col = _excel_find_col(df, ["Parkir", "Parking"])
    roof_col = _excel_find_col(df, ["Roof/Deck", "Roof Deck", "Roof"])
    mep_out_col = _excel_find_col(df, ["MEP Outdoor", "MEP Out", "Outdoor MEP"])
    koridor_col = _excel_find_col(df, ["Koridor/Lobby", "Koridor Lobby", "Corridor", "Lobby"])
    stair_col = _excel_find_col(df, ["Stair, MEP, Etc", "Stair MEP Etc", "Service", "Stair"])
    unit_col = _excel_find_col(df, ["Unit Area", "Unit", "NFA", "Retail"])
    office_col = _excel_find_col(df, ["Office"])

    gfa_col = _excel_find_col(df, ["GFA"])
    sgfa_col = _excel_find_col(df, ["SGFA", "SG FA"])
    nfa_col = _excel_find_col(df, ["NFA"])

    missing_required = []
    if not fl_col:
        missing_required.append("FL")
    if not space_col:
        missing_required.append("Space Type")
    if not unit_col:
        missing_required.append("Unit Area")

    if missing_required:
        raise ExcelImportError(
            'The "Area Input" sheet is missing required columns: '
            f'{", ".join(missing_required)}. Please use the latest downloaded Excel template.'
        )

    records = []

    for _, row in df.iterrows():
        fl = str(row.get(fl_col, "")).strip() if fl_col else ""

        if fl in ["", "-", "nan", "None"]:
            continue

        # Skip repeated header/total rows
        if fl.upper() in ["FL", "FLOOR", "TOTAL", "GRAND TOTAL"]:
            continue

        space_type = str(row.get(space_col, "")).strip() if space_col else ""

        parkir = _excel_safe_float(row.get(parkir_col, 0.0)) if parkir_col else 0.0
        roof = _excel_safe_float(row.get(roof_col, 0.0)) if roof_col else 0.0
        mep_out = _excel_safe_float(row.get(mep_out_col, 0.0)) if mep_out_col else 0.0

        koridor = _excel_safe_float(row.get(koridor_col, 0.0)) if koridor_col else 0.0
        stair = _excel_safe_float(row.get(stair_col, 0.0)) if stair_col else 0.0
        unit_area = _excel_safe_float(row.get(unit_col, 0.0)) if unit_col else 0.0
        office = _excel_safe_float(row.get(office_col, 0.0)) if office_col else 0.0

        detailed_values_exist = any(
            value > 0
            for value in [koridor, stair, unit_area, office]
        )

        if gfa_col and sgfa_col and nfa_col and not detailed_values_exist:
            gfa = _excel_safe_float(row.get(gfa_col, 0.0))
            sgfa = _excel_safe_float(row.get(sgfa_col, 0.0))
            nfa = _excel_safe_float(row.get(nfa_col, 0.0))

            if gfa > 0 or sgfa > 0 or nfa > 0:
                unit_area = nfa
                office = 0.0
                koridor = max(0.0, sgfa - nfa)
                stair = max(0.0, gfa - sgfa)

        records.append(
            {
                "FL": fl,
                "Space Type": space_type if space_type else "Unit",
                "Floor to Floor Height (m)": _excel_safe_float(row.get(f2f_col, 0.0)) if f2f_col else 0.0,
                "Typical Unit": _excel_safe_int(row.get(typical_col, 0)) if typical_col else 0,
                "Parkir": parkir,
                "Roof/Deck": roof,
                "MEP Outdoor": mep_out,
                "Koridor/Lobby": koridor,
                "Stair, MEP, Etc": stair,
                "Unit Area": unit_area,
                "Office": office,
            }
        )

    return clean_area_records(records)

def read_pintu_sheet(excel_bytes, area_df=None):
    try:
        raw = pd.read_excel(
            io.BytesIO(excel_bytes),
            sheet_name="Pintu",
            header=None,
            engine="openpyxl",
        )
    except Exception:
        return []

    # Generated Pintu sheet:
    # row 3 = headers, row 4 onward = data
    header_row_idx = None

    for i in range(min(10, len(raw))):
        row_norm = [_excel_norm_col(v) for v in raw.iloc[i].tolist()]
        if (
            _excel_norm_col("Pintu Kayu") in row_norm
            and _excel_norm_col("Pintu Besi") in row_norm
            and _excel_norm_col("Pintu Kaca") in row_norm
        ):
            header_row_idx = i
            break

    if header_row_idx is None:
        return []

    df = raw.iloc[header_row_idx + 1:].copy()
    df.columns = raw.iloc[header_row_idx].tolist()
    df = df.dropna(how="all").reset_index(drop=True)

    wood_col = _excel_find_col(df, ["Pintu Kayu", "Wood Door", "Wooden Door"])
    steel_col = _excel_find_col(df, ["Pintu Besi", "Steel Door"])
    glass_col = _excel_find_col(df, ["Pintu Kaca", "Glass Door"])

    if area_df is None or len(area_df) == 0:
        # Fallback, but linked formula columns may be blank if Excel did not recalculate.
        fl_col = _excel_find_col(df, ["FL", "Floor"])
        space_col = _excel_find_col(df, ["Space Type", "Floor Type", "Type"])
        f2f_col = _excel_find_col(df, ["Height (m)", "Floor to Floor Height (m)", "Height"])
        typical_col = _excel_find_col(df, ["Typical Unit"])

        records = []

        for _, row in df.iterrows():
            fl = str(row.get(fl_col, "")).strip() if fl_col else ""

            if fl in ["", "-", "nan", "None"] or fl.upper() in ["TOTAL", "FL"]:
                continue

            records.append(
                {
                    "FL": fl,
                    "Space Type": str(row.get(space_col, "")).strip() if space_col else "",
                    "Floor to Floor Height (m)": _excel_safe_float(row.get(f2f_col, 0.0)) if f2f_col else 0.0,
                    "Typical Unit": _excel_safe_int(row.get(typical_col, 0)) if typical_col else 0,
                    "Pintu Kayu": _excel_safe_int(row.get(wood_col, 0)) if wood_col else 0,
                    "Pintu Besi": _excel_safe_int(row.get(steel_col, 0)) if steel_col else 0,
                    "Pintu Kaca": _excel_safe_int(row.get(glass_col, 0)) if glass_col else 0,
                }
            )

        return clean_door_records(records)

    # Preferred path: use Area Input as row anchor.
    area_df = area_df.reset_index(drop=True).copy()
    records = []

    for i, area_row in area_df.iterrows():
        if i >= len(df):
            break

        door_row = df.iloc[i]

        records.append(
            {
                "FL": str(area_row.get("FL", "")).strip(),
                "Space Type": str(area_row.get("Space Type", "")).strip(),
                "Floor to Floor Height (m)": _excel_safe_float(
                    area_row.get("Floor to Floor Height (m)", 0.0)
                ),
                "Typical Unit": int(_excel_safe_float(area_row.get("Typical Unit", 0))),
                "Pintu Kayu": _excel_safe_int(door_row.get(wood_col, 0)) if wood_col else 0,
                "Pintu Besi": _excel_safe_int(door_row.get(steel_col, 0)) if steel_col else 0,
                "Pintu Kaca": _excel_safe_int(door_row.get(glass_col, 0)) if glass_col else 0,
            }
        )

    return clean_door_records(records)

def read_external_sheet(excel_bytes):
    try:
        df = _read_excel_sheet_with_header(
            excel_bytes,
            "Eksternal",
            header_candidates=["No", "Item", "Unit", "Qty", "Rate"],
        )
    except Exception:
        return [], {}

    no_col = _excel_find_col(df, ["No", "NO"])
    item_col = _excel_find_col(df, ["Item"])
    unit_col = _excel_find_col(df, ["Unit"])
    qty_col = _excel_find_col(df, ["Qty", "Quantity"])
    rate_col = _excel_find_col(df, ["Rate", "Harga"])

    other_records = []
    landscape_data = {}

    for _, row in df.iterrows():
        no = str(row.get(no_col, "")).strip() if no_col else ""
        item = str(row.get(item_col, "")).strip() if item_col else ""

        if item in ["", "-", "nan", "None"]:
            continue

        if no.upper() in ["NO", "TOTAL", "GRAND TOTAL"]:
            continue

        unit = str(row.get(unit_col, "")).strip() if unit_col else ""
        qty = _excel_safe_float(row.get(qty_col, 0.0)) if qty_col else 0.0
        rate = _excel_safe_float(row.get(rate_col, 0.0)) if rate_col else 0.0

        item_lower = item.lower()

        # -----------------------------
        # Landscape special rows
        # -----------------------------
        if item_lower == "landscape area":
            landscape_data["area_landscape_qty_calc"] = qty
            landscape_data["area_landscape_rate_calc"] = rate
            continue

        if item_lower == "hardscape %":
            landscape_data["area_hardscape_pct_calc"] = qty
            landscape_data["area_hardscape_rate_calc"] = rate
            continue

        if item_lower == "softscape %":
            landscape_data["area_softscape_pct_calc"] = qty
            landscape_data["area_softscape_rate_calc"] = rate
            continue

        # -----------------------------
        # Normal external works rows
        # -----------------------------
        other_records.append(
            {
                "No": no,
                "Item": item,
                "Unit": unit if unit else "ls",
                "Qty": qty,
                "Rate": rate,
            }
        )

    # Recalculate landscape amount from imported source values.
    landscape_qty = _excel_safe_float(
        landscape_data.get("area_landscape_qty_calc", 0.0)
    )
    hardscape_pct = _excel_safe_float(
        landscape_data.get("area_hardscape_pct_calc", 0.0)
    )
    softscape_pct = _excel_safe_float(
        landscape_data.get("area_softscape_pct_calc", 0.0)
    )
    hardscape_rate = _excel_safe_float(
        landscape_data.get("area_hardscape_rate_calc", 0.0)
    )
    softscape_rate = _excel_safe_float(
        landscape_data.get("area_softscape_rate_calc", 0.0)
    )

    hardscape_area = landscape_qty * hardscape_pct / 100.0
    softscape_area = landscape_qty * softscape_pct / 100.0

    hardscape_amount = hardscape_area * hardscape_rate
    softscape_amount = softscape_area * softscape_rate

    landscape_amount = hardscape_amount + softscape_amount
    landscape_rate = landscape_amount / landscape_qty if landscape_qty > 0 else 0.0

    landscape_data["area_hardscape_area_calc"] = hardscape_area
    landscape_data["area_softscape_area_calc"] = softscape_area
    landscape_data["area_landscape_amount_calc"] = landscape_amount
    landscape_data["area_landscape_rate_calc"] = landscape_rate

    return other_records, landscape_data

def read_residential_area_sheet(excel_bytes):
    try:
        df = _read_excel_sheet_with_header(
            excel_bytes,
            "Residential Area",
            header_candidates=["No", "Item", "Unit", "Qty", "Rate"],
        )
    except Exception:
        return []

    no_col = _excel_find_col(df, ["No", "NO"])
    item_col = _excel_find_col(df, ["Item"])
    unit_col = _excel_find_col(df, ["Unit"])
    qty_col = _excel_find_col(df, ["Qty", "Quantity"])
    rate_col = _excel_find_col(df, ["Rate", "Harga"])

    records = []

    for _, row in df.iterrows():
        item = str(row.get(item_col, "")).strip() if item_col else ""

        if item in ["", "-", "nan", "None"]:
            continue

        records.append(
            {
                "No": str(row.get(no_col, "")).strip() if no_col else "",
                "Item": item,
                "Unit": str(row.get(unit_col, "")).strip() if unit_col else "ls",
                "Qty": _excel_safe_float(row.get(qty_col, 0.0)) if qty_col else 0.0,
                "Rate": _excel_safe_float(row.get(rate_col, 0.0)) if rate_col else 0.0,
            }
        )

    return records


def _default_earthwork_detail_rows():
    return [
        {"code": "1.2.1", "description": "Cut Fill", "unit": "m2", "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "1.2.2", "description": "Dewatering", "unit": "ls", "quantity": 1.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "1.2.3", "description": "Soil Improvement", "unit": "m2", "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "1.2.4", "description": "Shoring System", "unit": "ls", "quantity": 1.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "1.2.5", "description": "Others", "unit": "ls", "quantity": 1.0, "unit_price": 0.0, "amount": 0.0},
    ]


def read_earthworks_sheet(excel_bytes):
    try:
        raw = pd.read_excel(
            io.BytesIO(excel_bytes),
            sheet_name="Earthworks",
            header=None,
            engine="openpyxl",
        )
    except ValueError:
        return None
    except Exception:
        return None

    header_row_idx = None

    for i in range(min(12, len(raw))):
        row_norm = [_excel_norm_col(v) for v in raw.iloc[i].tolist()]
        required_hits = sum(
            1 for key in ["Code", "Description", "Unit", "Quantity", "Unit Price (Rp)"]
            if _excel_norm_col(key) in row_norm
        )

        if required_hits >= 4:
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ExcelImportError(
            'The "Earthworks" sheet header could not be read. Earthworks import was skipped.'
        )

    df = raw.iloc[header_row_idx + 1:].copy()
    df.columns = raw.iloc[header_row_idx].tolist()
    df = df.loc[:, [str(c).strip() not in ["", "nan", "None"] for c in df.columns]]
    df = df.dropna(how="all").reset_index(drop=True)

    code_col = _excel_find_col(df, ["Code", "Item Code"])
    desc_col = _excel_find_col(df, ["Description", "Item"])
    unit_col = _excel_find_col(df, ["Unit"])
    qty_col = _excel_find_col(df, ["Quantity", "Qty"])
    unit_price_col = _excel_find_col(df, ["Unit Price (Rp)", "Unit Price", "Rate", "Harga"])

    missing = []
    if not code_col:
        missing.append("Code")
    if not desc_col:
        missing.append("Description")
    if not unit_col:
        missing.append("Unit")
    if not qty_col:
        missing.append("Quantity")
    if not unit_price_col:
        missing.append("Unit Price (Rp)")

    if missing:
        raise ExcelImportError(
            'The "Earthworks" sheet is missing required columns: '
            f'{", ".join(missing)}. Earthworks import was skipped.'
        )

    rows = []

    for _, row in df.iterrows():
        code = str(row.get(code_col, "")).strip()
        description = str(row.get(desc_col, "")).strip()

        if code in ["", "-", "nan", "None"] and description in ["", "-", "nan", "None"]:
            continue

        summary_labels = {
            "TOTAL",
            "GRAND TOTAL",
            "GBA",
            "EARTHWORKS DETAIL TOTAL",
            "DERIVED EARTHWORK PRICE",
        }
        if code.upper() in summary_labels or description.upper() in summary_labels:
            continue

        unit = str(row.get(unit_col, "")).strip().lower()
        unit = unit if unit in ["m2", "ls"] else "ls"

        raw_quantity = row.get(qty_col, "")
        quantity_is_blank = raw_quantity is None or str(raw_quantity).strip() in ["", "nan", "None"]
        quantity = 1.0 if unit == "ls" and quantity_is_blank else _excel_safe_float(raw_quantity, 0.0)

        unit_price = _excel_safe_float(row.get(unit_price_col, 0.0))
        amount = quantity * unit_price

        rows.append({
            "code": code,
            "description": description,
            "unit": unit,
            "quantity": quantity,
            "unit_price": unit_price,
            "amount": amount,
        })

    return rows


def _default_foundation_detail_rows():
    return [
        {"code": "1", "description": "Supply Tiang Pancang", "unit": "m'", "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "2", "description": "Install Tiang Pancang", "unit": "m'", "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
    ]


def read_foundation_sheet(excel_bytes):
    try:
        raw = pd.read_excel(
            io.BytesIO(excel_bytes),
            sheet_name="Foundation",
            header=None,
            engine="openpyxl",
        )
    except ValueError:
        return None
    except Exception:
        return None

    header_row_idx = None

    for i in range(min(12, len(raw))):
        row_norm = [_excel_norm_col(v) for v in raw.iloc[i].tolist()]
        required_hits = sum(
            1 for key in ["Code", "Description", "Unit", "Quantity", "Unit Price (Rp)"]
            if _excel_norm_col(key) in row_norm
        )

        if required_hits >= 4:
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ExcelImportError(
            'The "Foundation" sheet header could not be read. Foundation import was skipped.'
        )

    df = raw.iloc[header_row_idx + 1:].copy()
    df.columns = raw.iloc[header_row_idx].tolist()
    df = df.loc[:, [str(c).strip() not in ["", "nan", "None"] for c in df.columns]]
    df = df.dropna(how="all").reset_index(drop=True)

    code_col = _excel_find_col(df, ["Code", "Item Code"])
    desc_col = _excel_find_col(df, ["Description", "Item"])
    unit_col = _excel_find_col(df, ["Unit"])
    qty_col = _excel_find_col(df, ["Quantity", "Qty"])
    unit_price_col = _excel_find_col(df, ["Unit Price (Rp)", "Unit Price", "Rate", "Harga"])

    missing = []
    if not code_col:
        missing.append("Code")
    if not desc_col:
        missing.append("Description")
    if not unit_col:
        missing.append("Unit")
    if not qty_col:
        missing.append("Quantity")
    if not unit_price_col:
        missing.append("Unit Price (Rp)")

    if missing:
        raise ExcelImportError(
            'The "Foundation" sheet is missing required columns: '
            f'{", ".join(missing)}. Foundation import was skipped.'
        )

    by_code = {}
    summary_labels = {
        "TOTAL",
        "GRAND TOTAL",
        "GBA",
        "FOUNDATION DETAIL TOTAL",
        "DERIVED FOUNDATION RATE",
        "CURRENT FOUNDATION RATE",
        "DIFFERENCE",
    }

    for _, row in df.iterrows():
        code = str(row.get(code_col, "")).strip()
        description = str(row.get(desc_col, "")).strip()

        if code in ["", "-", "nan", "None"] and description in ["", "-", "nan", "None"]:
            continue

        if code.upper() in summary_labels or description.upper() in summary_labels:
            continue

        by_code[code] = {
            "code": code,
            "description": description,
            "unit": str(row.get(unit_col, "")).strip(),
            "quantity": _excel_safe_float(row.get(qty_col, 0.0)),
            "unit_price": _excel_safe_float(row.get(unit_price_col, 0.0)),
            "amount": 0.0,
        }

    rows = []
    for default_row in _default_foundation_detail_rows():
        imported = by_code.get(default_row["code"], {})
        row = {**default_row, **imported}
        row["code"] = default_row["code"]
        row["description"] = default_row["description"]
        row["unit"] = default_row["unit"]
        row["amount"] = _excel_safe_float(row.get("quantity", 0.0)) * _excel_safe_float(row.get("unit_price", 0.0))
        rows.append(row)

    return rows


def _default_structural_detail_rows():
    return [
        {"code": "1", "description": "Sub/Superstructure", "unit": "m3", "ratio": 0.0, "waste_factor": 0.0, "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "2", "description": "Bekisting", "unit": "m2", "ratio": 0.0, "waste_factor": 0.0, "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "3", "description": "Pembesian", "unit": "kg", "ratio": 0.0, "waste_factor": 0.0, "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "4", "description": "Readymix Concrete", "unit": "m3", "ratio": 0.0, "waste_factor": 0.0, "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "5", "description": "Rebar", "unit": "kg", "ratio": 0.0, "waste_factor": 0.0, "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "6", "description": "Prestress Works", "unit": "ls", "ratio": 0.0, "waste_factor": 0.0, "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "7", "description": "Steelworks", "unit": "kg", "ratio": 0.0, "waste_factor": 0.0, "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "8", "description": "Others", "unit": "m3", "ratio": 0.0, "waste_factor": 0.0, "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
    ]


def read_structural_sheet(excel_bytes):
    try:
        raw = pd.read_excel(
            io.BytesIO(excel_bytes),
            sheet_name="Structural",
            header=None,
            engine="openpyxl",
        )
    except ValueError:
        return None
    except Exception:
        return None

    header_row_idx = None

    for i in range(min(12, len(raw))):
        row_norm = [_excel_norm_col(v) for v in raw.iloc[i].tolist()]
        required_hits = sum(
            1 for key in ["Code", "Description", "Unit", "Ratio", "Quantity", "Unit Price (Rp)"]
            if _excel_norm_col(key) in row_norm
        )

        if required_hits >= 5:
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ExcelImportError(
            'The "Structural" sheet header could not be read. Structural import was skipped.'
        )

    df = raw.iloc[header_row_idx + 1:].copy()
    df.columns = raw.iloc[header_row_idx].tolist()
    df = df.loc[:, [str(c).strip() not in ["", "nan", "None"] for c in df.columns]]
    df = df.dropna(how="all").reset_index(drop=True)

    code_col = _excel_find_col(df, ["Code", "Item Code"])
    desc_col = _excel_find_col(df, ["Description", "Item"])
    unit_col = _excel_find_col(df, ["Unit"])
    ratio_col = _excel_find_col(df, ["Ratio"])
    waste_col = _excel_find_col(df, ["Waste Factor", "Waste"])
    qty_col = _excel_find_col(df, ["Quantity", "Qty"])
    unit_price_col = _excel_find_col(df, ["Unit Price (Rp)", "Unit Price", "Rate", "Harga"])

    missing = []
    if not code_col:
        missing.append("Code")
    if not desc_col:
        missing.append("Description")
    if not unit_col:
        missing.append("Unit")
    if not ratio_col:
        missing.append("Ratio")
    if not qty_col:
        missing.append("Quantity")
    if not unit_price_col:
        missing.append("Unit Price (Rp)")

    if missing:
        raise ExcelImportError(
            'The "Structural" sheet is missing required columns: '
            f'{", ".join(missing)}. Structural import was skipped.'
        )

    by_code = {}
    summary_labels = {
        "TOTAL",
        "GRAND TOTAL",
        "GBA",
        "STRUCTURAL DETAIL TOTAL",
        "DERIVED STRUCTURAL RATE",
        "CURRENT STRUCTURAL RATE",
        "DIFFERENCE",
    }

    for _, row in df.iterrows():
        code = str(row.get(code_col, "")).strip()
        description = str(row.get(desc_col, "")).strip()

        if code in ["", "-", "nan", "None"] and description in ["", "-", "nan", "None"]:
            continue

        if code.upper() in summary_labels or description.upper() in summary_labels:
            continue

        by_code[code] = {
            "code": code,
            "description": description,
            "unit": str(row.get(unit_col, "")).strip().lower(),
            "ratio": _excel_safe_float(row.get(ratio_col, 0.0)),
            "waste_factor": _excel_safe_float(row.get(waste_col, 0.0)) if waste_col else 0.0,
            "quantity": _excel_safe_float(row.get(qty_col, 0.0)),
            "unit_price": _excel_safe_float(row.get(unit_price_col, 0.0)),
            "amount": 0.0,
        }

    rows = []
    for default_row in _default_structural_detail_rows():
        imported = by_code.get(default_row["code"], {})
        row = {**default_row, **imported}
        row["code"] = default_row["code"]
        row["description"] = default_row["description"]
        row["unit"] = default_row["unit"]
        rows.append(row)

    return rows


def _default_architectural_detail_rows():
    return [
        {"code": "1", "description": "Basic Finishes Work", "unit": "m2", "factor": 0.0, "overlap": 0.0, "waste": 0.0, "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "2.1", "description": "Aluminium Facade / Window Wall", "unit": "m2", "factor": 0.0, "overlap": 0.0, "waste": 0.0, "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "2.2", "description": "Kisi2 Facade / Double Skin", "unit": "m2", "factor": 0.0, "overlap": 0.0, "waste": 0.0, "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "2.3", "description": "Precast Facade", "unit": "m2", "factor": 0.0, "overlap": 0.0, "waste": 0.0, "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "3", "description": "Pintu Kaca Dalam Ruangan", "unit": "unit", "factor": 0.0, "overlap": 0.0, "waste": 0.0, "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "4", "description": "Railing Balkon", "unit": "m'", "factor": 0.0, "overlap": 0.0, "waste": 0.0, "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "5", "description": "Pintu Kayu", "unit": "unit", "factor": 0.0, "overlap": 0.0, "waste": 0.0, "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "6", "description": "Pintu Besi", "unit": "unit", "factor": 0.0, "overlap": 0.0, "waste": 0.0, "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "7", "description": "Shower Screen", "unit": "unit", "factor": 0.0, "overlap": 0.0, "waste": 0.0, "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "8", "description": "Marble / Door Jamb Lift", "unit": "m'", "factor": 0.0, "overlap": 0.0, "waste": 0.0, "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "9", "description": "Interior - Main Lobby & Typical Lobby", "unit": "m2", "factor": 0.0, "overlap": 0.0, "waste": 0.0, "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "10", "description": "Signage / Fixtures", "unit": "ls", "factor": 0.0, "overlap": 0.0, "waste": 0.0, "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "11", "description": "Gondola", "unit": "unit", "factor": 0.0, "overlap": 0.0, "waste": 0.0, "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "12", "description": "Roof - Skylight", "unit": "m2", "factor": 0.0, "overlap": 0.0, "waste": 0.0, "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "13.1", "description": "Sanitary Fittings - T. Wanita", "unit": "unit", "factor": 0.0, "overlap": 0.0, "waste": 0.0, "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "13.2", "description": "Sanitary Fittings - T. Pria", "unit": "unit", "factor": 0.0, "overlap": 0.0, "waste": 0.0, "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "13.3", "description": "Sanitary Fittings - T. Disable", "unit": "unit", "factor": 0.0, "overlap": 0.0, "waste": 0.0, "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "13.4", "description": "Sanitary Fittings - Musholla", "unit": "unit", "factor": 0.0, "overlap": 0.0, "waste": 0.0, "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "13.5", "description": "Sanitary Fittings - Toilet Unit", "unit": "unit", "factor": 0.0, "overlap": 0.0, "waste": 0.0, "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "14", "description": "Kitchen Equipment", "unit": "unit", "factor": 0.0, "overlap": 0.0, "waste": 0.0, "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "15.1", "description": "Ironmongeries - Pintu Kayu", "unit": "unit", "factor": 0.0, "overlap": 0.0, "waste": 0.0, "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "15.2", "description": "Ironmongeries - Pintu Besi", "unit": "unit", "factor": 0.0, "overlap": 0.0, "waste": 0.0, "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "16.1", "description": "Keramik & HT", "unit": "m2", "factor": 0.0, "overlap": 0.0, "waste": 0.0, "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "16.2", "description": "Marmer", "unit": "m2", "factor": 0.0, "overlap": 0.0, "waste": 0.0, "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "16.3", "description": "Vinyl", "unit": "m2", "factor": 0.0, "overlap": 0.0, "waste": 0.0, "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "17", "description": "Carpet", "unit": "m2", "factor": 0.0, "overlap": 0.0, "waste": 0.0, "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
        {"code": "18", "description": "Kaca", "unit": "m2", "factor": 0.0, "overlap": 0.0, "waste": 0.0, "quantity": 0.0, "unit_price": 0.0, "amount": 0.0},
    ]


def read_architectural_sheet(excel_bytes):
    try:
        raw = pd.read_excel(
            io.BytesIO(excel_bytes),
            sheet_name="Architectural",
            header=None,
            engine="openpyxl",
        )
    except ValueError:
        return None
    except Exception:
        return None

    header_row_idx = None

    for i in range(min(12, len(raw))):
        row_norm = [_excel_norm_col(v) for v in raw.iloc[i].tolist()]
        required_hits = sum(
            1 for key in ["Code", "Description", "Unit", "Quantity", "Unit Price (Rp)"]
            if _excel_norm_col(key) in row_norm
        )

        if required_hits >= 4:
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ExcelImportError(
            'The "Architectural" sheet header could not be read. Architectural import was skipped.'
        )

    df = raw.iloc[header_row_idx + 1:].copy()
    df.columns = raw.iloc[header_row_idx].tolist()
    df = df.loc[:, [str(c).strip() not in ["", "nan", "None"] for c in df.columns]]
    df = df.dropna(how="all").reset_index(drop=True)

    code_col = _excel_find_col(df, ["Code", "Item Code"])
    desc_col = _excel_find_col(df, ["Description", "Item"])
    unit_col = _excel_find_col(df, ["Unit"])
    factor_col = _excel_find_col(df, ["Factor / %", "Factor", "Ratio", "Percentage", "%"])
    overlap_col = _excel_find_col(df, ["Overlap"])
    waste_col = _excel_find_col(df, ["Waste"])
    qty_col = _excel_find_col(df, ["Quantity", "Qty"])
    unit_price_col = _excel_find_col(df, ["Unit Price (Rp)", "Unit Price", "Rate", "Harga"])

    missing = []
    if not code_col:
        missing.append("Code")
    if not desc_col:
        missing.append("Description")
    if not unit_col:
        missing.append("Unit")
    if not qty_col:
        missing.append("Quantity")
    if not unit_price_col:
        missing.append("Unit Price (Rp)")

    if missing:
        raise ExcelImportError(
            'The "Architectural" sheet is missing required columns: '
            f'{", ".join(missing)}. Architectural import was skipped.'
        )

    by_code = {}
    summary_labels = {
        "TOTAL",
        "GRAND TOTAL",
        "GFA",
        "ARCHITECTURAL DETAIL TOTAL",
        "DERIVED ARCHITECTURAL RATE",
        "CURRENT ARCHITECTURAL RATE",
        "DIFFERENCE",
    }

    for _, row in df.iterrows():
        code = str(row.get(code_col, "")).strip()
        description = str(row.get(desc_col, "")).strip()

        if code in ["", "-", "nan", "None"] and description in ["", "-", "nan", "None"]:
            continue

        if code.upper() in summary_labels or description.upper() in summary_labels:
            continue

        by_code[code] = {
            "code": code,
            "description": description,
            "unit": str(row.get(unit_col, "")).strip(),
            "factor": _excel_safe_float(row.get(factor_col, 0.0)) if factor_col else 0.0,
            "overlap": _excel_safe_float(row.get(overlap_col, 0.0)) if overlap_col else 0.0,
            "waste": _excel_safe_float(row.get(waste_col, 0.0)) if waste_col else 0.0,
            "quantity": _excel_safe_float(row.get(qty_col, 0.0)),
            "unit_price": _excel_safe_float(row.get(unit_price_col, 0.0)),
            "amount": 0.0,
        }

    rows = []
    for default_row in _default_architectural_detail_rows():
        imported = by_code.get(default_row["code"], {})
        row = {**default_row, **imported}
        row["code"] = default_row["code"]
        row["description"] = default_row["description"]
        row["unit"] = default_row["unit"]
        row["amount"] = _excel_safe_float(row.get("quantity", 0.0)) * _excel_safe_float(row.get("unit_price", 0.0))
        rows.append(row)

    return rows


def create_area_excel_form_bytes(
    project_name="",
    upper_floors=5,
    basements=1,
    include_roof_machine=True,
    include_roof=True,
    architectural_detail_rows=None,
    earthwork_detail_rows=None,
    foundation_detail_rows=None,
    structural_detail_rows=None,
    architectural_base_values=None,
    earthwork_gba=0.0,
    foundation_gba=0.0,
    structural_gba=0.0,
):
    output = io.BytesIO()

    wb = Workbook()
    ws = wb.active
    ws.title = "Area Input"

    # ==================================================
    # BASIC STYLES
    # ==================================================
    dark = "111827"
    blue = "D9EAF7"
    green = "E2F0D9"
    gray = "E7E6E6"
    yellow = "FFF2CC"
    white = "FFFFFF"
    formula_fill = "F3F4F6"

    thin_gray = Side(style="thin", color="BFBFBF")
    medium_dark = Side(style="medium", color=dark)

    border_all = Border(
        left=thin_gray,
        right=thin_gray,
        top=thin_gray,
        bottom=thin_gray,
    )

    def style_range(ws_, cell_range, fill=None, font_color=dark, bold=False, size=10):
        for row in ws_[cell_range]:
            for cell in row:
                if fill:
                    cell.fill = PatternFill("solid", fgColor=fill)
                cell.font = Font(color=font_color, bold=bold, size=size)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border_all

    def unlock_range(ws_, cell_range):
        for row in ws_[cell_range]:
            for cell in row:
                cell.protection = Protection(locked=False)

    def lock_range(ws_, cell_range):
        for row in ws_[cell_range]:
            for cell in row:
                cell.protection = Protection(locked=True)

    # ==================================================
    # FLOOR ROW GENERATION
    # ==================================================
    floor_rows = []

    if include_roof_machine:
        floor_rows.append(("Roof Machine", "Roof", 3.0))

    if include_roof:
        floor_rows.append(("Roof", "Roof", 3.0))

    for i in range(int(upper_floors), 1, -1):
        floor_rows.append((f"{i}F", "Unit", 3.2))

    floor_rows.append(("1F", "Lobby", 4.5))

    for i in range(1, int(basements) + 1):
        fl_name = "LG" if i == 1 else f"B{i - 1}"
        floor_rows.append((fl_name, "Carpark", 3.2))

    start_row = 5
    end_row = start_row + len(floor_rows) - 1
    total_row = end_row + 1

    # ==================================================
    # AREA INPUT SHEET
    # ==================================================
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:N1")
    ws["A1"] = f"AREA INPUT FORM - {project_name or 'PROJECT'}"
    ws["A1"].font = Font(bold=True, color=white, size=14)
    ws["A1"].fill = PatternFill("solid", fgColor=dark)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:D2")
    ws["A2"] = "BASIC FLOOR INFO"

    ws.merge_cells("E2:G2")
    ws["E2"] = "GBA"

    ws.merge_cells("H2:H2")
    ws["H2"] = "GFA"

    ws.merge_cells("I2:I2")
    ws["I2"] = "SGFA"

    ws.merge_cells("J2:K2")
    ws["J2"] = "NFA"

    ws.merge_cells("L2:N2")
    ws["L2"] = "FORMULA CHECK"

    ws["A3"] = "FL"
    ws["B3"] = "Space Type"
    ws["C3"] = "Height (m)"
    ws["D3"] = "Typical Unit"
    ws["E3"] = "Parkir"
    ws["F3"] = "Roof/Deck"
    ws["G3"] = "MEP Outdoor"
    ws["H3"] = "Stair, MEP, Etc"
    ws["I3"] = "Koridor/Lobby"
    ws["J3"] = "Unit Area"
    ws["K3"] = "Office"
    ws["L3"] = "TOTAL"
    ws["M3"] = "GBA"
    ws["N3"] = "GFA"

    # Hidden stable import key row
    ws["A4"] = "FL"
    ws["B4"] = "Space Type"
    ws["C4"] = "Floor to Floor Height (m)"
    ws["D4"] = "Typical Unit"
    ws["E4"] = "Parkir"
    ws["F4"] = "Roof/Deck"
    ws["G4"] = "MEP Outdoor"
    ws["H4"] = "Stair, MEP, Etc"
    ws["I4"] = "Koridor/Lobby"
    ws["J4"] = "Unit Area"
    ws["K4"] = "Office"
    ws["L4"] = "TOTAL"
    ws["M4"] = "GBA"
    ws["N4"] = "GFA"
    ws.row_dimensions[4].hidden = True

    style_range(ws, "A2:D2", blue, bold=True)
    style_range(ws, "E2:G2", gray, bold=True)
    style_range(ws, "H2:H2", yellow, bold=True)
    style_range(ws, "I2:I2", green, bold=True)
    style_range(ws, "J2:K2", blue, bold=True)
    style_range(ws, "L2:N2", formula_fill, bold=True)

    style_range(ws, "A3:N4", dark, font_color=white, bold=True)

    for r, (fl, space_type, height) in enumerate(floor_rows, start=start_row):
        ws.cell(r, 1).value = fl
        ws.cell(r, 2).value = space_type
        ws.cell(r, 3).value = height
        ws.cell(r, 4).value = 0

        for c in range(5, 12):
            ws.cell(r, c).value = 0

        ws.cell(r, 12).value = f"=SUM(E{r}:K{r})"
        ws.cell(r, 13).value = f"=L{r}"
        ws.cell(r, 14).value = f"=L{r}-SUM(E{r}:G{r})"

    ws.cell(total_row, 1).value = "TOTAL"
    ws.cell(total_row, 1).font = Font(bold=True)

    for c in range(3, 15):
        col = get_column_letter(c)
        ws.cell(total_row, c).value = f"=SUM({col}{start_row}:{col}{end_row})"
        ws.cell(total_row, c).font = Font(bold=True)

    style_range(ws, f"A{start_row}:N{total_row}", None)
    style_range(ws, f"L{start_row}:N{total_row}", formula_fill)
    style_range(ws, f"A{total_row}:N{total_row}", dark, font_color=white, bold=True)

    # Editable cells only
    lock_range(ws, f"A1:N{total_row}")
    unlock_range(ws, f"A{start_row}:K{end_row}")

    # Data validation for Space Type
    dv_space = DataValidation(
        type="list",
        formula1='"Roof,Unit,Lobby,Ramp,Carpark,Facility,Office"',
        allow_blank=False,
    )
    ws.add_data_validation(dv_space)
    dv_space.add(f"B{start_row}:B{end_row}")

    # Number formats
    for row in ws.iter_rows(min_row=start_row, max_row=total_row, min_col=3, max_col=14):
        for cell in row:
            cell.number_format = '#,##0.00'

    for row in ws.iter_rows(min_row=start_row, max_row=total_row, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = '#,##0'

    for col, width in {
        "A": 16,
        "B": 16,
        "C": 12,
        "D": 12,
        "E": 14,
        "F": 14,
        "G": 14,
        "H": 16,
        "I": 16,
        "J": 14,
        "K": 12,
        "L": 14,
        "M": 14,
        "N": 14,
    }.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A5"
    ws.protection.sheet = True
    ws.protection.password = "area"

    # ==================================================
    # AREA CHART SHEET
    # ==================================================
    ws_chart = wb.create_sheet("Area Chart")
    ws_chart.sheet_view.showGridLines = False

    ws_chart.merge_cells("A1:I1")
    ws_chart["A1"] = "BUILDING AREA SECTION CHART"
    ws_chart["A1"].font = Font(bold=True, color=white, size=14)
    ws_chart["A1"].fill = PatternFill("solid", fgColor=dark)
    ws_chart["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_chart.row_dimensions[1].height = 26

    chart_header_row = 3
    chart_data_start_row = 4
    chart_data_end_row = chart_data_start_row + len(floor_rows) - 1

    chart_headers = [
        "FL",
        "Spacer",
        "Office",
        "Unit Area",
        "Stair, MEP, Etc",
        "Koridor/Lobby",
        "Parkir",
        "MEP Outdoor",
        "Roof/Deck",
    ]

    for c, h in enumerate(chart_headers, start=1):
        ws_chart.cell(chart_header_row, c).value = h

    style_range(
        ws_chart,
        f"A{chart_header_row}:I{chart_header_row}",
        dark,
        font_color=white,
        bold=True,
    )

    for idx, _floor in enumerate(floor_rows):
        r = chart_data_start_row + idx
        area_row = start_row + idx

        ws_chart.cell(r, 1).value = f"='Area Input'!A{area_row}"

        # Centering spacer
        ws_chart.cell(r, 2).value = (
            f"=(MAX('Area Input'!$M${start_row}:$M${end_row})-'Area Input'!M{area_row})/2"
        )

        ws_chart.cell(r, 3).value = f"='Area Input'!K{area_row}"  # Office
        ws_chart.cell(r, 4).value = f"='Area Input'!J{area_row}"  # Unit Area
        ws_chart.cell(r, 5).value = f"='Area Input'!H{area_row}"  # Stair, MEP, Etc
        ws_chart.cell(r, 6).value = f"='Area Input'!I{area_row}"  # Koridor/Lobby
        ws_chart.cell(r, 7).value = f"='Area Input'!E{area_row}"  # Parkir
        ws_chart.cell(r, 8).value = f"='Area Input'!G{area_row}"  # MEP Outdoor
        ws_chart.cell(r, 9).value = f"='Area Input'!F{area_row}"  # Roof/Deck

    style_range(ws_chart, f"A{chart_data_start_row}:I{chart_data_end_row}", None)

    for row in ws_chart.iter_rows(
        min_row=chart_data_start_row,
        max_row=chart_data_end_row,
        min_col=2,
        max_col=9,
    ):
        for cell in row:
            cell.number_format = '#,##0.00'

    # Create horizontal stacked bar chart
    area_chart = BarChart()
    area_chart.type = "bar"
    area_chart.grouping = "stacked"
    area_chart.overlap = 100
    area_chart.title = "Building Area Section"
    area_chart.y_axis.title = "Floor"
    area_chart.x_axis.title = "Area (m2)"
    area_chart.height = 18
    area_chart.width = 34

    data = Reference(
        ws_chart,
        min_col=2,
        max_col=9,
        min_row=chart_header_row,
        max_row=chart_data_end_row,
    )

    categories = Reference(
        ws_chart,
        min_col=1,
        min_row=chart_data_start_row,
        max_row=chart_data_end_row,
    )

    area_chart.add_data(data, titles_from_data=True)
    area_chart.set_categories(categories)

    # Hide spacer visually
    try:
        area_chart.series[0].graphicalProperties.noFill = True
        area_chart.series[0].graphicalProperties.line.noFill = True
    except Exception:
        pass

    area_chart.y_axis.scaling.orientation = "maxMin"

    ws_chart.add_chart(area_chart, "K3")

    # Make helper table visible for now so we can confirm chart source works.
    for col, width in {
        "A": 16,
        "B": 12,
        "C": 14,
        "D": 14,
        "E": 18,
        "F": 16,
        "G": 14,
        "H": 14,
        "I": 14,
        "K": 4,
    }.items():
        ws_chart.column_dimensions[col].width = width

    ws_chart.freeze_panes = "A4"

    # ==================================================
    # PINTU SHEET
    # ==================================================
    ws_door = wb.create_sheet("Pintu")
    ws_door.sheet_view.showGridLines = False

    door_headers = [
        "FL",
        "Space Type",
        "Height (m)",
        "Typical Unit",
        "Pintu Kayu",
        "Pintu Besi",
        "Pintu Kaca",
    ]

    ws_door.merge_cells("A1:G1")
    ws_door["A1"] = "PINTU INPUT FORM"
    ws_door["A1"].font = Font(bold=True, color=white, size=14)
    ws_door["A1"].fill = PatternFill("solid", fgColor=dark)
    ws_door["A1"].alignment = Alignment(horizontal="center", vertical="center")

    for c, h in enumerate(door_headers, start=1):
        ws_door.cell(3, c).value = h

    style_range(ws_door, "A3:G3", dark, font_color=white, bold=True)

    for idx, (fl, space_type, height) in enumerate(floor_rows):
        r = 4 + idx
        area_row = start_row + idx

        # Linked to Area Input sheet
        ws_door.cell(r, 1).value = f"='Area Input'!A{area_row}"
        ws_door.cell(r, 2).value = f"='Area Input'!B{area_row}"
        ws_door.cell(r, 3).value = f"='Area Input'!C{area_row}"
        ws_door.cell(r, 4).value = f"='Area Input'!D{area_row}"

        # User input
        ws_door.cell(r, 5).value = 0
        ws_door.cell(r, 6).value = 0
        ws_door.cell(r, 7).value = 0

    door_total_row = 4 + len(floor_rows)
    ws_door.cell(door_total_row, 1).value = "TOTAL"
    for c in range(4, 8):
        col = get_column_letter(c)
        ws_door.cell(door_total_row, c).value = f"=SUM({col}4:{col}{door_total_row - 1})"

    style_range(ws_door, f"A4:G{door_total_row}", None)
    style_range(ws_door, f"A{door_total_row}:G{door_total_row}", dark, font_color=white, bold=True)

    lock_range(ws_door, f"A1:G{door_total_row}")
    unlock_range(ws_door, f"E4:G{door_total_row - 1}")

    for col in range(1, 8):
        ws_door.column_dimensions[get_column_letter(col)].width = 18

    ws_door.freeze_panes = "A4"
    ws_door.protection.sheet = True
    ws_door.protection.password = "area"

    # ==================================================
    # EKSTERNAL SHEET
    # ==================================================
    ws_ext = wb.create_sheet("Eksternal")
    ws_ext.sheet_view.showGridLines = False

    ws_ext.merge_cells("A1:F1")
    ws_ext["A1"] = "EKSTERNAL INPUT FORM"
    ws_ext["A1"].font = Font(bold=True, color=white, size=14)
    ws_ext["A1"].fill = PatternFill("solid", fgColor=dark)
    ws_ext["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ext_headers = ["No", "Item", "Unit", "Qty", "Rate", "Amount"]
    for c, h in enumerate(ext_headers, start=1):
        ws_ext.cell(3, c).value = h

    ext_rows = [
        ["1", "Landscape Area", "m2", 0, 0],
        ["1.1", "Hardscape %", "%", 0, 0],
        ["1.2", "Softscape %", "%", 0, 0],
        ["2", "SBO : PJU", "tk", 0, 0],
        ["3", "Drainage System", "m2", 0, 0],
        ["4", "Boundary Wall & Gates", "m1", 0, 0],
        ["5", "Infrastructure-Access road", "m2", 0, 0],
        ["6", "Others", "ls", 0, 0],
    ]

    for r, row in enumerate(ext_rows, start=4):
        for c, val in enumerate(row, start=1):
            ws_ext.cell(r, c).value = val

        no_value = str(row[0]).strip()

        if no_value == "1":
            # Landscape Area row:
            # Amount = Landscape Area x Landscape Rate
            # Landscape Rate is calculated from Hardscape + Softscape amount.
            ws_ext.cell(r, 6).value = "=D4*E4"

        elif no_value == "1.1":
            # Hardscape Amount = Landscape Area x Hardscape % x Hardscape Rate
            ws_ext.cell(r, 6).value = "=D4*(D5/100)*E5"

        elif no_value == "1.2":
            # Softscape Amount = Landscape Area x Softscape % x Softscape Rate
            ws_ext.cell(r, 6).value = "=D4*(D6/100)*E6"

        else:
            ws_ext.cell(r, 6).value = f"=D{r}*E{r}"

    # Landscape blended rate:
    # Landscape Rate = (Hardscape Amount + Softscape Amount) / Landscape Area
    ws_ext["E4"] = '=IF(D4>0,(F5+F6)/D4,0)'

    ext_total_row = 4 + len(ext_rows)
    ws_ext.cell(ext_total_row, 1).value = "TOTAL"
    ws_ext.cell(ext_total_row, 6).value = f"=SUM(F4:F{ext_total_row - 1})"

    style_range(ws_ext, "A3:F3", dark, font_color=white, bold=True)
    style_range(ws_ext, f"A4:F{ext_total_row}", None)
    style_range(ws_ext, f"A{ext_total_row}:F{ext_total_row}", dark, font_color=white, bold=True)
    style_range(ws_ext, f"F4:F{ext_total_row}", formula_fill)

    lock_range(ws_ext, f"A1:F{ext_total_row}")
    unlock_range(ws_ext, f"A4:E{ext_total_row - 1}")

    # Lock calculated landscape blended rate
    ws_ext["E4"].protection = Protection(locked=True)

    for col, width in {
        "A": 10,
        "B": 34,
        "C": 12,
        "D": 14,
        "E": 16,
        "F": 18,
    }.items():
        ws_ext.column_dimensions[col].width = width

    for row in ws_ext.iter_rows(min_row=4, max_row=ext_total_row, min_col=4, max_col=6):
        for cell in row:
            cell.number_format = '#,##0.00'

    ws_ext.freeze_panes = "A4"
    ws_ext.protection.sheet = True
    ws_ext.protection.password = "area"

    # ==================================================
    # RESIDENTIAL AREA SHEET
    # ==================================================
    ws_res = wb.create_sheet("Residential Area")
    ws_res.sheet_view.showGridLines = False

    ws_res.merge_cells("A1:F1")
    ws_res["A1"] = "RESIDENTIAL FACILITY INPUT FORM"
    ws_res["A1"].font = Font(bold=True, color=white, size=14)
    ws_res["A1"].fill = PatternFill("solid", fgColor=dark)
    ws_res["A1"].alignment = Alignment(horizontal="center", vertical="center")

    res_headers = ["No", "Item", "Unit", "Qty", "Rate", "Amount"]
    for c, h in enumerate(res_headers, start=1):
        ws_res.cell(3, c).value = h

    res_rows = [
        ["1", "Swimming Pool", "m2", 0, 0],
        ["2", "Club House / Fitness Centre", "ls", 0, 0],
        ["3", "Pool Deck", "m2", 0, 0],
    ]

    for r, row in enumerate(res_rows, start=4):
        for c, val in enumerate(row, start=1):
            ws_res.cell(r, c).value = val
        ws_res.cell(r, 6).value = f"=D{r}*E{r}"

    res_total_row = 4 + len(res_rows)
    ws_res.cell(res_total_row, 1).value = "TOTAL"
    ws_res.cell(res_total_row, 6).value = f"=SUM(F4:F{res_total_row - 1})"

    style_range(ws_res, "A3:F3", dark, font_color=white, bold=True)
    style_range(ws_res, f"A4:F{res_total_row}", None)
    style_range(ws_res, f"A{res_total_row}:F{res_total_row}", dark, font_color=white, bold=True)
    style_range(ws_res, f"F4:F{res_total_row}", formula_fill)

    lock_range(ws_res, f"A1:F{res_total_row}")
    unlock_range(ws_res, f"A4:E{res_total_row - 1}")

    for col, width in {
        "A": 10,
        "B": 34,
        "C": 12,
        "D": 14,
        "E": 16,
        "F": 18,
    }.items():
        ws_res.column_dimensions[col].width = width

    for row in ws_res.iter_rows(min_row=4, max_row=res_total_row, min_col=4, max_col=6):
        for cell in row:
            cell.number_format = '#,##0.00'

    ws_res.freeze_panes = "A4"
    ws_res.protection.sheet = True
    ws_res.protection.password = "area"

    # ==================================================
    # EARTHWORKS SHEET
    # ==================================================
    ws_earth = wb.create_sheet("Earthworks")
    ws_earth.sheet_view.showGridLines = False

    ws_earth.merge_cells("A1:F1")
    ws_earth["A1"] = "EARTHWORKS DETAIL BREAKDOWN"
    ws_earth["A1"].font = Font(bold=True, color=white, size=14)
    ws_earth["A1"].fill = PatternFill("solid", fgColor=dark)
    ws_earth["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws_earth.merge_cells("A2:F2")
    ws_earth["A2"] = "Preview only: this Earthworks breakdown does not control Cost Analysis yet."
    ws_earth["A2"].font = Font(italic=True, color=dark, size=10)
    ws_earth["A2"].alignment = Alignment(horizontal="left", vertical="center")

    earth_headers = ["Code", "Description", "Unit", "Quantity", "Unit Price (Rp)", "Amount (Rp)"]
    for c, h in enumerate(earth_headers, start=1):
        ws_earth.cell(3, c).value = h

    style_range(ws_earth, "A3:F3", dark, font_color=white, bold=True)

    earth_rows = earthwork_detail_rows if isinstance(earthwork_detail_rows, list) and earthwork_detail_rows else _default_earthwork_detail_rows()

    for r, row in enumerate(earth_rows, start=4):
        row = row if isinstance(row, dict) else {}
        unit = str(row.get("unit", "ls") or "ls").strip().lower()
        unit = unit if unit in ["m2", "ls"] else "ls"
        quantity = _excel_safe_float(row.get("quantity", 1.0 if unit == "ls" else 0.0))
        unit_price = _excel_safe_float(row.get("unit_price", 0.0))

        ws_earth.cell(r, 1).value = str(row.get("code", "")).strip()
        ws_earth.cell(r, 2).value = str(row.get("description", "")).strip()
        ws_earth.cell(r, 3).value = unit
        ws_earth.cell(r, 4).value = quantity
        ws_earth.cell(r, 5).value = unit_price
        ws_earth.cell(r, 6).value = f"=D{r}*E{r}"

    earth_total_row = 4 + len(earth_rows)
    ws_earth.cell(earth_total_row, 1).value = "TOTAL"
    ws_earth.cell(earth_total_row, 6).value = f"=SUM(F4:F{earth_total_row - 1})"

    summary_start = earth_total_row + 2
    ws_earth.cell(summary_start, 1).value = "GBA"
    ws_earth.cell(summary_start, 2).value = _excel_safe_float(earthwork_gba)
    ws_earth.cell(summary_start + 1, 1).value = "Earthworks Detail Total"
    ws_earth.cell(summary_start + 1, 2).value = f"=F{earth_total_row}"
    ws_earth.cell(summary_start + 2, 1).value = "Derived Earthwork Price"
    ws_earth.cell(summary_start + 2, 2).value = f"=IF(B{summary_start}>0,B{summary_start + 1}/B{summary_start},0)"

    style_range(ws_earth, f"A4:F{earth_total_row}", None)
    style_range(ws_earth, f"A{earth_total_row}:F{earth_total_row}", dark, font_color=white, bold=True)
    style_range(ws_earth, f"A{summary_start}:B{summary_start + 2}", formula_fill, bold=True)
    style_range(ws_earth, f"F4:F{earth_total_row}", formula_fill)

    lock_range(ws_earth, f"A1:F{summary_start + 2}")
    unlock_range(ws_earth, f"A4:E{earth_total_row - 1}")

    dv_earth_unit = DataValidation(
        type="list",
        formula1='"m2,ls"',
        allow_blank=False,
    )
    ws_earth.add_data_validation(dv_earth_unit)
    dv_earth_unit.add(f"C4:C{earth_total_row - 1}")

    for col, width in {
        "A": 12,
        "B": 28,
        "C": 12,
        "D": 14,
        "E": 18,
        "F": 18,
    }.items():
        ws_earth.column_dimensions[col].width = width

    for row in ws_earth.iter_rows(min_row=4, max_row=summary_start + 2, min_col=4, max_col=6):
        for cell in row:
            cell.number_format = '#,##0.00'

    for cell in [ws_earth.cell(summary_start, 2), ws_earth.cell(summary_start + 1, 2), ws_earth.cell(summary_start + 2, 2)]:
        cell.number_format = '#,##0.00'

    ws_earth.freeze_panes = "A4"
    ws_earth.protection.sheet = True
    ws_earth.protection.password = "area"

    # ==================================================
    # FOUNDATION SHEET
    # ==================================================
    ws_found = wb.create_sheet("Foundation")
    ws_found.sheet_view.showGridLines = False

    ws_found.merge_cells("A1:F1")
    ws_found["A1"] = "FOUNDATION DETAIL BREAKDOWN"
    ws_found["A1"].font = Font(bold=True, color=white, size=14)
    ws_found["A1"].fill = PatternFill("solid", fgColor=dark)
    ws_found["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws_found.merge_cells("A2:F2")
    ws_found["A2"] = "Detail-derived rate only. Cost Analysis changes only after Apply Foundation Detail Rate in the app."
    ws_found["A2"].font = Font(italic=True, color=dark, size=10)
    ws_found["A2"].alignment = Alignment(horizontal="left", vertical="center")

    foundation_headers = ["Code", "Description", "Unit", "Quantity", "Unit Price (Rp)", "Amount (Rp)"]
    for c, h in enumerate(foundation_headers, start=1):
        ws_found.cell(3, c).value = h

    style_range(ws_found, "A3:F3", dark, font_color=white, bold=True)

    foundation_rows = foundation_detail_rows if isinstance(foundation_detail_rows, list) and foundation_detail_rows else _default_foundation_detail_rows()
    foundation_defaults = _default_foundation_detail_rows()

    for idx, default_row in enumerate(foundation_defaults):
        r = 4 + idx
        row = foundation_rows[idx] if idx < len(foundation_rows) and isinstance(foundation_rows[idx], dict) else {}
        quantity = _excel_safe_float(row.get("quantity", 0.0))
        unit_price = _excel_safe_float(row.get("unit_price", 0.0))

        ws_found.cell(r, 1).value = default_row["code"]
        ws_found.cell(r, 2).value = default_row["description"]
        ws_found.cell(r, 3).value = default_row["unit"]
        ws_found.cell(r, 4).value = quantity
        ws_found.cell(r, 5).value = unit_price
        ws_found.cell(r, 6).value = f"=D{r}*E{r}"

    foundation_total_row = 4 + len(foundation_defaults)
    ws_found.cell(foundation_total_row, 1).value = "TOTAL"
    ws_found.cell(foundation_total_row, 6).value = f"=SUM(F4:F{foundation_total_row - 1})"

    foundation_summary_start = foundation_total_row + 2
    ws_found.cell(foundation_summary_start, 1).value = "GBA"
    ws_found.cell(foundation_summary_start, 2).value = _excel_safe_float(foundation_gba)
    ws_found.cell(foundation_summary_start + 1, 1).value = "Foundation Detail Total"
    ws_found.cell(foundation_summary_start + 1, 2).value = f"=F{foundation_total_row}"
    ws_found.cell(foundation_summary_start + 2, 1).value = "Derived Foundation Rate"
    ws_found.cell(foundation_summary_start + 2, 2).value = f"=IF(B{foundation_summary_start}>0,B{foundation_summary_start + 1}/B{foundation_summary_start},0)"

    style_range(ws_found, f"A4:F{foundation_total_row}", None)
    style_range(ws_found, f"A{foundation_total_row}:F{foundation_total_row}", dark, font_color=white, bold=True)
    style_range(ws_found, f"A{foundation_summary_start}:B{foundation_summary_start + 2}", formula_fill, bold=True)
    style_range(ws_found, f"F4:F{foundation_total_row}", formula_fill)

    lock_range(ws_found, f"A1:F{foundation_summary_start + 2}")
    unlock_range(ws_found, f"D4:E{foundation_total_row - 1}")

    for col, width in {
        "A": 12,
        "B": 28,
        "C": 12,
        "D": 14,
        "E": 18,
        "F": 18,
    }.items():
        ws_found.column_dimensions[col].width = width

    for row in ws_found.iter_rows(min_row=4, max_row=foundation_summary_start + 2, min_col=4, max_col=6):
        for cell in row:
            cell.number_format = '#,##0.00'

    for cell in [
        ws_found.cell(foundation_summary_start, 2),
        ws_found.cell(foundation_summary_start + 1, 2),
        ws_found.cell(foundation_summary_start + 2, 2),
    ]:
        cell.number_format = '#,##0.00'

    ws_found.freeze_panes = "A4"
    ws_found.protection.sheet = True
    ws_found.protection.password = "area"

    # ==================================================
    # STRUCTURAL SHEET
    # ==================================================
    ws_struct = wb.create_sheet("Structural")
    ws_struct.sheet_view.showGridLines = False

    ws_struct.merge_cells("A1:H1")
    ws_struct["A1"] = "STRUCTURAL DETAIL BREAKDOWN"
    ws_struct["A1"].font = Font(bold=True, color=white, size=14)
    ws_struct["A1"].fill = PatternFill("solid", fgColor=dark)
    ws_struct["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws_struct.merge_cells("A2:H2")
    ws_struct["A2"] = "Detail-derived rate only. Cost Analysis changes only after Apply Structural Detail Rate in the app."
    ws_struct["A2"].font = Font(italic=True, color=dark, size=10)
    ws_struct["A2"].alignment = Alignment(horizontal="left", vertical="center")

    structural_headers = [
        "Code",
        "Description",
        "Unit",
        "Ratio",
        "Waste Factor",
        "Quantity",
        "Unit Price (Rp)",
        "Amount (Rp)",
    ]
    for c, h in enumerate(structural_headers, start=1):
        ws_struct.cell(3, c).value = h

    style_range(ws_struct, "A3:H3", dark, font_color=white, bold=True)

    struct_rows = structural_detail_rows if isinstance(structural_detail_rows, list) and structural_detail_rows else _default_structural_detail_rows()
    struct_defaults = _default_structural_detail_rows()
    struct_start_row = 4

    for idx, default_row in enumerate(struct_defaults):
        r = struct_start_row + idx
        row = struct_rows[idx] if idx < len(struct_rows) and isinstance(struct_rows[idx], dict) else {}

        ws_struct.cell(r, 1).value = default_row["code"]
        ws_struct.cell(r, 2).value = default_row["description"]
        ws_struct.cell(r, 3).value = default_row["unit"]
        ws_struct.cell(r, 4).value = _excel_safe_float(row.get("ratio", 0.0))
        ws_struct.cell(r, 5).value = _excel_safe_float(row.get("waste_factor", 0.0))
        ws_struct.cell(r, 7).value = _excel_safe_float(row.get("unit_price", 0.0))

        if idx == 0:
            ws_struct.cell(r, 6).value = f"={_excel_safe_float(structural_gba)}*D{r}"
        elif idx in [1, 2, 3]:
            ws_struct.cell(r, 6).value = f"=F{struct_start_row}*D{r}"
        elif idx == 4:
            ws_struct.cell(r, 6).value = f"=F{struct_start_row}*D{r}*E{r}"
        elif idx in [5, 6]:
            ws_struct.cell(r, 6).value = _excel_safe_float(row.get("quantity", 0.0))
        else:
            ws_struct.cell(r, 6).value = f"=F{struct_start_row}"

        ws_struct.cell(r, 8).value = f"=F{r}*G{r}"

    struct_total_row = struct_start_row + len(struct_defaults)
    ws_struct.cell(struct_total_row, 1).value = "TOTAL"
    ws_struct.cell(struct_total_row, 8).value = f"=SUM(H{struct_start_row}:H{struct_total_row - 1})"

    struct_summary_start = struct_total_row + 2
    ws_struct.cell(struct_summary_start, 1).value = "GBA"
    ws_struct.cell(struct_summary_start, 2).value = _excel_safe_float(structural_gba)
    ws_struct.cell(struct_summary_start + 1, 1).value = "Structural Detail Total"
    ws_struct.cell(struct_summary_start + 1, 2).value = f"=H{struct_total_row}"
    ws_struct.cell(struct_summary_start + 2, 1).value = "Derived Structural Rate"
    ws_struct.cell(struct_summary_start + 2, 2).value = f"=IF(B{struct_summary_start}>0,B{struct_summary_start + 1}/B{struct_summary_start},0)"

    style_range(ws_struct, f"A{struct_start_row}:H{struct_total_row}", None)
    style_range(ws_struct, f"A{struct_total_row}:H{struct_total_row}", dark, font_color=white, bold=True)
    style_range(ws_struct, f"A{struct_summary_start}:B{struct_summary_start + 2}", formula_fill, bold=True)
    style_range(ws_struct, f"F{struct_start_row}:F{struct_total_row}", formula_fill)
    style_range(ws_struct, f"H{struct_start_row}:H{struct_total_row}", formula_fill)

    lock_range(ws_struct, f"A1:H{struct_summary_start + 2}")
    unlock_range(ws_struct, f"D{struct_start_row}:D{struct_start_row + 4}")
    unlock_range(ws_struct, f"E{struct_start_row + 4}:E{struct_start_row + 4}")
    unlock_range(ws_struct, f"F{struct_start_row + 5}:F{struct_start_row + 6}")
    unlock_range(ws_struct, f"G{struct_start_row}:G{struct_total_row - 1}")

    for col, width in {
        "A": 12,
        "B": 28,
        "C": 12,
        "D": 14,
        "E": 14,
        "F": 14,
        "G": 18,
        "H": 18,
    }.items():
        ws_struct.column_dimensions[col].width = width

    for row in ws_struct.iter_rows(min_row=struct_start_row, max_row=struct_summary_start + 2, min_col=4, max_col=8):
        for cell in row:
            cell.number_format = '#,##0.00'

    for cell in [
        ws_struct.cell(struct_summary_start, 2),
        ws_struct.cell(struct_summary_start + 1, 2),
        ws_struct.cell(struct_summary_start + 2, 2),
    ]:
        cell.number_format = '#,##0.00'

    ws_struct.freeze_panes = "A4"
    ws_struct.protection.sheet = True
    ws_struct.protection.password = "area"

    # ==================================================
    # ARCHITECTURAL SHEET
    # ==================================================
    ws_arch = wb.create_sheet("Architectural")
    ws_arch.sheet_view.showGridLines = False

    arch_base = architectural_base_values if isinstance(architectural_base_values, dict) else {}
    arch_gfa = _excel_safe_float(arch_base.get("gfa", 0.0))
    arch_facade = _excel_safe_float(arch_base.get("facade", 0.0))
    arch_rooms = _excel_safe_float(arch_base.get("rooms", 0.0))
    arch_glass_door = _excel_safe_float(arch_base.get("glass_door", 0.0))
    arch_wooden_door = _excel_safe_float(arch_base.get("wooden_door", 0.0))
    arch_steel_door = _excel_safe_float(arch_base.get("steel_door", 0.0))
    arch_lobby = _excel_safe_float(arch_base.get("lobby", 0.0))

    ws_arch.merge_cells("A1:I1")
    ws_arch["A1"] = "ARCHITECTURAL DETAIL BREAKDOWN"
    ws_arch["A1"].font = Font(bold=True, color=white, size=14)
    ws_arch["A1"].fill = PatternFill("solid", fgColor=dark)
    ws_arch["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws_arch.merge_cells("A2:I2")
    ws_arch["A2"] = "Detail-derived rate only. Cost Analysis changes only after Apply Architectural Detail Rate in the app."
    ws_arch["A2"].font = Font(italic=True, color=dark, size=10)
    ws_arch["A2"].alignment = Alignment(horizontal="left", vertical="center")

    arch_headers = [
        "Code",
        "Description",
        "Unit",
        "Factor / %",
        "Overlap",
        "Waste",
        "Quantity",
        "Unit Price (Rp)",
        "Amount (Rp)",
    ]
    for c, h in enumerate(arch_headers, start=1):
        ws_arch.cell(3, c).value = h

    style_range(ws_arch, "A3:I3", dark, font_color=white, bold=True)

    arch_rows = architectural_detail_rows if isinstance(architectural_detail_rows, list) and architectural_detail_rows else _default_architectural_detail_rows()
    arch_defaults = _default_architectural_detail_rows()
    arch_start_row = 4

    for idx, default_row in enumerate(arch_defaults):
        r = arch_start_row + idx
        row = arch_rows[idx] if idx < len(arch_rows) and isinstance(arch_rows[idx], dict) else {}
        code = default_row["code"]

        ws_arch.cell(r, 1).value = code
        ws_arch.cell(r, 2).value = default_row["description"]
        ws_arch.cell(r, 3).value = default_row["unit"]
        ws_arch.cell(r, 4).value = _excel_safe_float(row.get("factor", 0.0))
        ws_arch.cell(r, 5).value = _excel_safe_float(row.get("overlap", 0.0))
        ws_arch.cell(r, 6).value = _excel_safe_float(row.get("waste", 0.0))
        ws_arch.cell(r, 8).value = _excel_safe_float(row.get("unit_price", 0.0))

        if code == "1":
            ws_arch.cell(r, 7).value = arch_gfa
        elif code in ["2.1", "2.2", "2.3"]:
            ws_arch.cell(r, 7).value = f"={arch_facade}*(D{r}/100)"
        elif code == "3":
            ws_arch.cell(r, 7).value = arch_glass_door
        elif code == "4":
            ws_arch.cell(r, 7).value = f"={arch_rooms}*D{r}"
        elif code == "5":
            ws_arch.cell(r, 7).value = arch_wooden_door
        elif code == "6":
            ws_arch.cell(r, 7).value = arch_steel_door
        elif code in ["7", "8", "10", "11", "12", "13.1", "13.2", "13.3", "13.4", "17", "18"]:
            ws_arch.cell(r, 7).value = _excel_safe_float(row.get("quantity", 0.0))
        elif code == "9":
            ws_arch.cell(r, 7).value = arch_lobby
        elif code == "13.5":
            ws_arch.cell(r, 7).value = f"={arch_rooms}*D{r}"
        elif code == "14":
            ws_arch.cell(r, 7).value = arch_rooms
        elif code == "15.1":
            ws_arch.cell(r, 7).value = arch_wooden_door
        elif code == "15.2":
            ws_arch.cell(r, 7).value = arch_steel_door
        elif code in ["16.1", "16.2", "16.3"]:
            ws_arch.cell(r, 7).value = f"={arch_gfa}*(D{r}/100)*E{r}*F{r}"
        else:
            ws_arch.cell(r, 7).value = _excel_safe_float(row.get("quantity", 0.0))

        ws_arch.cell(r, 9).value = f"=G{r}*H{r}"

    arch_total_row = arch_start_row + len(arch_defaults)
    ws_arch.cell(arch_total_row, 1).value = "TOTAL"
    ws_arch.cell(arch_total_row, 9).value = f"=SUM(I{arch_start_row}:I{arch_total_row - 1})"

    arch_summary_start = arch_total_row + 2
    ws_arch.cell(arch_summary_start, 1).value = "GFA"
    ws_arch.cell(arch_summary_start, 2).value = arch_gfa
    ws_arch.cell(arch_summary_start + 1, 1).value = "Architectural Detail Total"
    ws_arch.cell(arch_summary_start + 1, 2).value = f"=I{arch_total_row}"
    ws_arch.cell(arch_summary_start + 2, 1).value = "Derived Architectural Rate"
    ws_arch.cell(arch_summary_start + 2, 2).value = f"=IF(B{arch_summary_start}>0,B{arch_summary_start + 1}/B{arch_summary_start},0)"

    style_range(ws_arch, f"A{arch_start_row}:I{arch_total_row}", None)
    style_range(ws_arch, f"A{arch_total_row}:I{arch_total_row}", dark, font_color=white, bold=True)
    style_range(ws_arch, f"A{arch_summary_start}:B{arch_summary_start + 2}", formula_fill, bold=True)
    style_range(ws_arch, f"G{arch_start_row}:G{arch_total_row}", formula_fill)
    style_range(ws_arch, f"I{arch_start_row}:I{arch_total_row}", formula_fill)

    lock_range(ws_arch, f"A1:I{arch_summary_start + 2}")
    unlock_range(ws_arch, f"D{arch_start_row}:F{arch_total_row - 1}")
    unlock_range(ws_arch, f"H{arch_start_row}:H{arch_total_row - 1}")
    for r in range(arch_start_row, arch_total_row):
        code = str(ws_arch.cell(r, 1).value)
        if code in ["7", "8", "10", "11", "12", "13.1", "13.2", "13.3", "13.4", "17", "18"]:
            ws_arch.cell(r, 7).protection = Protection(locked=False)

    for col, width in {
        "A": 12,
        "B": 36,
        "C": 12,
        "D": 14,
        "E": 14,
        "F": 14,
        "G": 14,
        "H": 18,
        "I": 18,
    }.items():
        ws_arch.column_dimensions[col].width = width

    for row in ws_arch.iter_rows(min_row=arch_start_row, max_row=arch_summary_start + 2, min_col=4, max_col=9):
        for cell in row:
            cell.number_format = '#,##0.00'

    for cell in [
        ws_arch.cell(arch_summary_start, 2),
        ws_arch.cell(arch_summary_start + 1, 2),
        ws_arch.cell(arch_summary_start + 2, 2),
    ]:
        cell.number_format = '#,##0.00'

    ws_arch.freeze_panes = "A4"
    ws_arch.protection.sheet = True
    ws_arch.protection.password = "area"

    # ==================================================
    # IMPORT GUIDE
    # ==================================================
    ws_guide = wb.create_sheet("Import Guide")
    guide_rows = [
        ["Sheet", "Purpose", "Editable Columns"],
        ["Area Input", "Main floor area input", "Editable source columns: FL, Space Type, Height, Typical Unit, Parkir, Roof/Deck, MEP Outdoor, Stair/MEP/Etc, Koridor/Lobby, Unit Area, Office"],
        ["Pintu", "Door quantity input", "Floor, Space Type, Height, and Typical Unit are linked from Area Input. Edit only Pintu Kayu, Pintu Besi, Pintu Kaca."],
        ["Eksternal", "External works input", "No, Item, Unit, Qty, Rate"],
        ["Residential Area", "Residential facility input", "No, Item, Unit, Qty, Rate"],
        ["Earthworks", "Earthworks detail preview input", "Code, Description, Unit, Quantity, Unit Price (Rp)"],
        ["Foundation", "Foundation detail-derived rate input", "Code, Description, Unit, Quantity, Unit Price (Rp)"],
        ["Structural", "Structural detail-derived rate input", "Ratio, Waste Factor for Rebar, Quantity for Prestress Works and Steelworks, Unit Price (Rp)"],
        ["Architectural", "Architectural detail-derived rate input", "Factor / %, Overlap, Waste, manual Quantity rows, Unit Price (Rp)"],
    ]

    for r, row in enumerate(guide_rows, start=1):
        for c, val in enumerate(row, start=1):
            ws_guide.cell(r, c).value = val

    style_range(ws_guide, "A1:C1", dark, font_color=white, bold=True)

    for col, width in {"A": 22, "B": 32, "C": 90}.items():
        ws_guide.column_dimensions[col].width = width

    ws_guide.protection.sheet = True
    ws_guide.protection.password = "area"

    wb.save(output)
    output.seek(0)

    return output.getvalue()
