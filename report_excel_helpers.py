import io
import re

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

from cost_calculation_helpers import build_cost_raw_from_project_data, calculate_live_costs
from project_database import get_project_type_data


def _safe_float(v, default=0.0):
    if v is None:
        return default

    try:
        import pandas as pd
        if pd.isna(v):
            return default
    except Exception:
        pass

    if isinstance(v, str):
        v = (
            v.strip()
            .replace("Rp", "")
            .replace("rp", "")
            .replace(",", "")
            .replace("%", "")
        )

        if v == "" or v.lower() in ["none", "nan", "null", "-"]:
            return default

    try:
        return float(v)
    except Exception:
        return default


def generate_exact_portfolio_excel(port_meta, port_data, port_assumptions):
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio Summary"

    # --- 1. Styling Definitions ---
    blue_fill = PatternFill(start_color="005A9C", end_color="005A9C", fill_type="solid")
    gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    
    white_font = Font(color="FFFFFF", bold=True, name='Calibri', size=11)
    bold_font = Font(bold=True, name='Calibri', size=10)
    reg_font = Font(bold=False, name='Calibri', size=10)
    small_font = Font(name='Calibri', size=9)
    
    black_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                          top=Side(style='thin'), bottom=Side(style='thin'))
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')

    # --- 2. Blue Header Section ---
    headers = [
        ["ASG GROUP PROPERTY DEVELOPMENT", f"VERSION : {port_meta.get('version', '')}"],
        ["QS & PROCUREMENT DIVISION", f"UPDATED : {port_meta.get('updated', '')}"],
        [port_meta.get('title', ''), f"CREATED : {port_meta.get('created', '')}"],
        [port_meta.get('ref', ''), ""]
    ]

    for r_idx, (text_left, text_right) in enumerate(headers, 1):
        for c_idx in range(1, 12):
            c = ws.cell(row=r_idx, column=c_idx)
            c.fill = blue_fill
            c.font = white_font
            if c_idx == 1: c.value = text_left
            if c_idx == 11: 
                c.value = text_right
                c.alignment = Alignment(horizontal='right')
        ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=10)

    # --- 3. Table Header (Rows 6 & 7) ---
    ws.merge_cells(start_row=6, start_column=1, end_row=7, end_column=1) # SN
    ws.merge_cells(start_row=6, start_column=2, end_row=7, end_column=2) # AREA
    ws.merge_cells(start_row=6, start_column=3, end_row=6, end_column=5) # BLDG AREA
    ws.merge_cells(start_row=6, start_column=6, end_row=6, end_column=7) # UNIT
    ws.merge_cells(start_row=6, start_column=8, end_row=7, end_column=8) # BUDGET
    ws.merge_cells(start_row=6, start_column=9, end_row=6, end_column=11) # COST RATIO

    header_labels = {
        (6, 1): "SN", (6, 2): "AREA", (6, 3): "BUILDING AREA (M2)", 
        (6, 6): "UNIT", (6, 8): "BUDGET ESTIMATE\nRP", (6, 9): "COST RATIO RP/M2",
        (7, 3): "GBA", (7, 4): "GFA", (7, 5): "SGFA", (7, 9): "GBA", (7, 10): "GFA", (7, 11): "SGFA"
    }

    for (r, c), text in header_labels.items():
        cell = ws.cell(row=r, column=c, value=text)
        cell.alignment = center_align
        cell.font = bold_font
        cell.fill = gray_fill

    for r in range(6, 8):
        for c in range(1, 12):
            ws.cell(row=r, column=c).border = black_border

    # --- 4. Data Rows ---
    current_row = 8
    for p_row in port_data:
        is_total = p_row.get("AREA") == "TOTAL"
        is_parking = "PARKING" in str(p_row.get("AREA", "")).upper()
        
        # Set SN and Area
        ws.cell(row=current_row, column=1, value=p_row.get("SN", "")).alignment = center_align
        
        area_cell = ws.cell(row=current_row, column=2, value=p_row.get("AREA", ""))
        area_cell.alignment = Alignment(horizontal='left', vertical='top' if is_parking else 'center', wrap_text=True)
        
        # Set Values
        cols = ["GBA", "GFA", "SGFA", "QTY", "UNIT", "BUDGET", "R_GBA", "R_GFA", "R_SGFA"]
        for i, key in enumerate(cols, 3):
            val = p_row.get(key, "")
            cell = ws.cell(row=current_row, column=i, value=val)
            cell.alignment = Alignment(vertical='top' if is_parking else 'center', horizontal='right' if i != 7 else 'center')
            
            if key in ["GBA", "GFA", "SGFA"]: cell.number_format = "#,##0.00"
            if key in ["BUDGET", "R_GBA", "R_GFA", "R_SGFA"]: cell.number_format = "#,##0"

        # Apply Row Styles
        for c in range(1, 12):
            ws.cell(row=current_row, column=c).border = black_border
            ws.cell(row=current_row, column=c).font = bold_font if (p_row.get("SN") or is_total) else reg_font
            if is_total: ws.cell(row=current_row, column=c).fill = gray_fill

        if is_total:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
            ws.cell(row=current_row, column=6, value="TOTAL").alignment = center_align

        # Adjust height for multi-line rows (Parking)
        if is_parking:
            ws.row_dimensions[current_row].height = 100 

        current_row += 1

    # --- 5. Assumptions Section ---
    current_row += 1
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=11)
    cell_assump = ws.cell(row=current_row, column=1, value="I.  ASSUMPTIONS")
    cell_assump.font = bold_font
    for c in range(1, 12):
        ws.cell(row=current_row, column=c).fill = yellow_fill
        ws.cell(row=current_row, column=c).border = black_border

    current_row += 1
    for _, a_row in port_assumptions.iterrows():
        ws.cell(row=current_row, column=1, value=a_row.get("No.", "")).alignment = center_align
        
        desc_cell = ws.cell(row=current_row, column=2, value=a_row.get("Assumption Description", ""))
        desc_cell.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=11)
        
        for c in range(1, 12):
            ws.cell(row=current_row, column=c).border = black_border
        current_row += 1

    # Column Widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    for c in ['C', 'D', 'E', 'H', 'I', 'J', 'K']:
        ws.column_dimensions[c].width = 16

    wb.save(output)
    return output.getvalue()


def normalize_header_token(value, trailing_dot=False):
    text = str(value or "").strip().upper()
    text = re.sub(r"(?<=[A-Z])\s+(?=\d)", "", text)
    text = re.sub(r"\s+", ".", text)
    text = re.sub(r"\.+", ".", text)

    if trailing_dot and text and not text.endswith("."):
        text = f"{text}."

    return text


def build_portfolio_meta_from_inputs(header_inputs):
    project_location = normalize_header_token(
        header_inputs.get("project_location", ""),
        trailing_dot=True,
    )
    project_name = normalize_header_token(header_inputs.get("project_name", ""))
    option_number = str(header_inputs.get("option_number", "")).strip()
    revision_number = str(header_inputs.get("revision_number", "")).strip()
    drawing_date = str(header_inputs.get("drawing_date", "")).strip()
    updated_date = str(header_inputs.get("updated_date", "")).strip()
    created_date = str(header_inputs.get("created_date", "")).strip()

    project_label = f"{project_location}{project_name}".strip(".")
    title_parts = [project_label] if project_label else []
    if option_number:
        title_parts.append(f"OPT.{option_number}")
    title_revision = "R(1)" if revision_number else ""
    if title_revision:
        title_parts.append(title_revision)
    title = "PROJECT PORTFOLIO"
    if title_parts:
        title = f"{title} | {' '.join(title_parts)}"

    ref_parts = ["REF. DATA"]
    if revision_number:
        ref_parts.append(f"R({revision_number})")
    ref = f"{' '.join(ref_parts)} | CONCEPT DWG {drawing_date}.DPA"

    version_parts = []
    if title_revision:
        version_parts.append("R (1)")
    if option_number:
        version_parts.append(f"OPT{option_number}")
    version = " ".join(version_parts) if version_parts else "-"

    return {
        "title": title,
        "ref": ref,
        "version": version,
        "updated": updated_date,
        "created": created_date,
    }

def get_recap_values(pdata):
    d = pdata.get("data", {})
    curr_type = pdata.get("type", "Hotel")
    pt_data = get_project_type_data(curr_type)
    raw = build_cost_raw_from_project_data(d, pt_data)
    costs = calculate_live_costs(raw)
    
    return {
        "EARTHWORKS": costs["t_earth"], "FOUNDATIONS": costs["t_found"], "STRUCTURAL WORKS": costs["t_struc"],
        "ARCHITECTURAL WORKS": costs["group_arch"], "FF & E": costs["group_ffe"], "M.E.P WORKS": costs["t_mep"],
        "UTILITY CONNECTION": costs["t_utility"], "EXTERNAL WORKS": costs["t_external"], "FACILITY": costs["group_misc"],
        "PRELIMINARIES WORKS": costs["t_preliminary"], "CONTINGENCIES": costs["t_contingency"], "HARDCOST": costs["grand_total_hc"],
        "CONSULTANCY SERVICES FEE": costs["t_consultancy"], "QS SERVICES": costs["t_qs"],
        "PROJECT MANAGEMENT SERVICES": costs["t_pm"], "INSURANCE COVERAGE": costs["t_insurance"],
        "SOFTCOST": costs["total_soft_cost"], "TOTAL, EXCLD PPN": costs["grand_total_project"]
    }

def generate_recap_excel(port_meta, projects):
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Recap Cost"

    

    # --- Generate Global Totals ---
    tot_cache = {}; global_cost = {}; tot_gba = tot_gfa = tot_sgfa = 0
    for pid, pdata in projects.items():
        vals = get_recap_values(pdata)
        tot_cache[pid] = vals
        for k, v in vals.items(): global_cost[k] = global_cost.get(k, 0) + v
        d = pdata.get("data", {})
        tot_gba += _safe_float(d.get("m_gba", 0)); tot_gfa += _safe_float(d.get("m_gfa", 0)); tot_sgfa += _safe_float(d.get("m_sgfa", 0))

    # Calculate global % divisors
    global_hc = global_cost.get("HARDCOST", 0)
    global_sc = global_cost.get("SOFTCOST", 0)
    safe_hc = global_hc if global_hc > 0 else 1
    safe_sc = global_sc if global_sc > 0 else 1

    # --- 1. Styling Definitions ---
    blue_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    white_font, bold_font, reg_font = Font(color="FFFFFF", bold=True, size=10), Font(bold=True, size=9), Font(size=9)
    black_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align, left_align = Alignment(horizontal='center', vertical='center', wrap_text=True), Alignment(horizontal='left', vertical='center')

    # --- 2. Blue Metadata Header ---
    headers = [["ASG GROUP PROPERTY DEVELOPMENT", f"VERSION : {port_meta.get('version', '')}"], ["QS & PROCUREMENT DIVISION", f"UPDATED : {port_meta.get('updated', '')}"], [port_meta.get('title', ''), f"CREATED : {port_meta.get('created', '')}"], [port_meta.get('ref', ''), ""]]
    for r_idx, (text_left, text_right) in enumerate(headers, 1):
        for c in range(1, 100): ws.cell(row=r_idx, column=c).fill = blue_fill
        ws.cell(row=r_idx, column=1, value=text_left).font = white_font
        ws.cell(row=r_idx, column=15, value=text_right).font = white_font
        ws.cell(row=r_idx, column=15).alignment = Alignment(horizontal='right', vertical='center')

    # --- 3. Static Table Headers ---
    static_cols = [("SN", 1), ("DESCRIPTION", 2), ("COA", 3), ("%", 4)]
    for name, col_idx in static_cols:
        ws.merge_cells(start_row=6, start_column=col_idx, end_row=9, end_column=col_idx)
        c = ws.cell(row=6, column=col_idx, value=name)
        c.alignment, c.font, c.fill = center_align, bold_font, PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for r in range(6, 10): ws.cell(row=r, column=col_idx).border = black_border

    # --- 4. Dynamic Project Headers ---
    bg_colors = ["EAEAEA", "FCE4D6", "F2DCDB", "E1D5E7", "DDEBF7", "E2EFDA", "D9E1F2", "F4B084", "FFF2CC"]
    project_list = [("TOTAL", {"name": "TOTAL"})] + list(projects.items())
    current_col = 5
    
    for i, (pid, pdata) in enumerate(project_list):
        group_fill = PatternFill(start_color=bg_colors[i % len(bg_colors)], end_color=bg_colors[i % len(bg_colors)], fill_type="solid")
        ws.merge_cells(start_row=6, start_column=current_col, end_row=6, end_column=current_col+4)
        c = ws.cell(row=6, column=current_col, value=pdata.get('name', 'PROJECT').upper())
        c.alignment, c.font, c.fill = center_align, bold_font, group_fill
        
        ws.cell(row=7, column=current_col, value="ESTIMATE").alignment = center_align
        ws.merge_cells(start_row=7, start_column=current_col+1, end_row=7, end_column=current_col+4)
        ws.cell(row=7, column=current_col+1, value="Cost Ratio (Rp/m2)").alignment = center_align
        
        for j, lbl in enumerate(["TOTAL", "GBA", "GFA", "SGFA", "NFA"]):
            ws.cell(row=8, column=current_col+j, value=lbl).alignment = center_align
            
        if pid == "TOTAL":
            gba, gfa, sgfa, nfa = tot_gba, tot_gfa, tot_sgfa, tot_gfa * 0.82
        else:
            d = pdata.get("data", {})
            gba, gfa, sgfa = _safe_float(d.get("m_gba", 0)), _safe_float(d.get("m_gfa", 0)), _safe_float(d.get("m_sgfa", 0))
            nfa = gfa * 0.82
        
        gba_f = gba if gba > 0 else 1; gfa_f = gfa if gfa > 0 else 1; sgfa_f = sgfa if sgfa > 0 else 1; nfa_f = nfa if nfa > 0 else 1
            
        for j, val in enumerate(["Rp", gba, gfa, sgfa, nfa]):
            c = ws.cell(row=9, column=current_col+j, value=val)
            c.alignment = center_align
            if j > 0: c.number_format = '#,##0'

        for r in range(6, 10):
            for c_idx in range(current_col, current_col+5):
                ws.cell(row=r, column=c_idx).fill = group_fill
                ws.cell(row=r, column=c_idx).border = black_border
        current_col += 5

    # --- 5. Data Rows Map (With Categories for %) ---
    row_mapping = [
        ("I", "HARDCOST", "118-14-000", True, "HC"), ("1", "PRELIMINARIES WORKS", "118-14-100", False, "HC"), 
        ("2", "EARTHWORKS", "118-14-200", False, "HC"), ("3", "FOUNDATIONS", "118-14-300", False, "HC"), 
        ("4", "STRUCTURAL WORKS", "118-14-500", False, "HC"), ("5", "ARCHITECTURAL WORKS", "118-14-600", False, "HC"),
        ("6", "FF & E", "118-14-700", False, "HC"), ("7", "M.E.P WORKS", "118-14-800", False, "HC"), 
        ("8", "UTILITY CONNECTION", "118-13-900", False, "HC"), ("9", "EXTERNAL WORKS", "118-14-930", False, "HC"), 
        ("10", "FACILITY", "118-14-960", False, "HC"), ("11", "CONTINGENCIES", "", False, "HC"),
        ("II", "SOFTCOST", "118-13-000", True, "SC_TOTAL"), ("1", "CONSULTANCY SERVICES FEE", "118-13-202", False, "SC"), 
        ("2", "QS SERVICES", "118-13-201", False, "SC"), ("3", "PROJECT MANAGEMENT SERVICES", "118-13-203", False, "SC"), 
        ("4", "INSURANCE COVERAGE", "118-13-300", False, "SC"), ("IV", "TOTAL, EXCLD PPN", "", True, "TOTAL")
    ]

    r_idx = 10
    for sn, desc, coa, is_bold, cat in row_mapping:
        # Calculate % based on Global Totals and Category
        global_val = global_cost.get(desc, 0)
        if cat == "HC": pct = global_val / safe_hc
        elif cat == "SC": pct = global_val / safe_sc
        elif cat == "SC_TOTAL": pct = global_val / safe_hc # Softcost total vs Hardcost
        elif cat == "TOTAL": pct = global_val / safe_hc # Grand total vs Hardcost
        else: pct = 0

        ws.cell(row=r_idx, column=1, value=sn).alignment = center_align
        ws.cell(row=r_idx, column=2, value=desc).alignment = left_align
        ws.cell(row=r_idx, column=3, value=coa).alignment = center_align
        
        # Write Percentage
        pct_cell = ws.cell(row=r_idx, column=4, value=pct)
        pct_cell.alignment = center_align
        pct_cell.number_format = '0.00%'

        for c in range(1, 5):
            ws.cell(row=r_idx, column=c).border = black_border
            ws.cell(row=r_idx, column=c).font = bold_font if is_bold else reg_font

        col_offset = 5
        for pid, pdata in project_list:
            val = global_cost.get(desc, 0) if pid == "TOTAL" else tot_cache[pid].get(desc, 0)
            
            if pid == "TOTAL":
                gba_f = tot_gba if tot_gba > 0 else 1; gfa_f = tot_gfa if tot_gfa > 0 else 1
                sgfa_f = tot_sgfa if tot_sgfa > 0 else 1; nfa_f = (tot_gfa*0.82) if tot_gfa > 0 else 1
            else:
                d = pdata.get("data", {})
                gba_f = _safe_float(d.get("m_gba", 1) if d.get("m_gba", 0) > 0 else 1)
                gfa_f = _safe_float(d.get("m_gfa", 1) if d.get("m_gfa", 0) > 0 else 1)
                sgfa_f = _safe_float(d.get("m_sgfa", 1) if d.get("m_sgfa", 0) > 0 else 1)
                nfa_f = gfa_f * 0.82
            
            for j, v in enumerate([val, val/gba_f, val/gfa_f, val/sgfa_f, val/nfa_f]):
                c = ws.cell(row=r_idx, column=col_offset+j, value=v)
                c.number_format = '#,##0'
                c.border = black_border
                c.font = bold_font if is_bold else reg_font
                if is_bold: c.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            col_offset += 5
        r_idx += 1

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    for c in range(5, col_offset): ws.column_dimensions[get_column_letter(c)].width = 16
    ws.freeze_panes = 'E10'
    wb.save(output)
    return output.getvalue()

